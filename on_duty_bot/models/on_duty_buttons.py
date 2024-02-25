import datetime
import openpyxl


F_NAME = "on_duty_bot/models/schedule.xlsx"  # Path to schedule file


def get_current_shift(schedule_file_name=F_NAME):
    """
    Parses work schedule and returns string of engeniers which are on duty this time.
    """

    cur_hour = int(datetime.datetime.now().strftime("%H"))
    cur_day = int(datetime.datetime.now().strftime("%d"))
    cur_month = int(datetime.datetime.now().strftime("%m"))
    cur_year = int(datetime.datetime.now().strftime("%Y"))

    day_shift_names = []
    night_shift_names = []

    book = openpyxl.load_workbook(schedule_file_name)  # open file
    sheet = book.active  # take active worksheet

    if 8 <= cur_hour and cur_hour < 20:  # find day shift names (I)
        for row in sheet.iter_rows(max_col=1):
            if row[0].value == "№ п/п":
                day_row_ind = row[0].row + 1
        for cell in sheet[day_row_ind]:
            if cell.value == cur_day:
                day_col_ind = cell.column
        for cell in sheet.iter_rows(min_col=day_col_ind, max_col=day_col_ind):
            if cell[0].value == "I":
                n_row = cell[0].row
                day_shift_names.append(sheet["C" + str(n_row - 1)].value)
        return (
            f"Сейчас на смене - {str(cur_day) +'.'+ str(cur_month).zfill(2) + '.' + str(cur_year)}"
            + "\n"
            + "\n".join(day_shift_names)
        )
    else:  # find night shift names (II)
        for row in sheet.iter_rows(max_col=1):
            if row[0].value == "№ п/п":
                day_row_ind = row[0].row + 1
        if 20 <= cur_hour:  #  if current time before 00:00
            for cell in sheet[day_row_ind]:
                if cell.value == cur_day:
                    day_col_ind = cell.column
        else:  # if current time after 00:00
            for cell in sheet[day_row_ind]:
                if cell.value == cur_day - 1:
                    day_col_ind = cell.column
        for cell in sheet.iter_rows(min_col=day_col_ind, max_col=day_col_ind):
            if cell[0].value == "II":
                n_row = cell[0].row
                night_shift_names.append(sheet["C" + str(n_row - 1)].value)
        return (
            f"Сейчас на смене - {str(cur_day) +'.'+ str(cur_month).zfill(2) + '.' + str(cur_year)}"
            + "\n"
            + "\n".join(night_shift_names)
        )


def get_previous_shift(schedule_file_name=F_NAME):
    """
    Parses work schedule and returns string of engeniers which were on duty previous shift.
    """

    cur_hour = int(datetime.datetime.now().strftime("%H"))
    cur_day = int(datetime.datetime.now().strftime("%d"))
    cur_month = int(datetime.datetime.now().strftime("%m"))
    cur_year = int(datetime.datetime.now().strftime("%Y"))

    day_shift_names = []
    night_shift_names = []

    book = openpyxl.load_workbook(schedule_file_name)  # open file
    sheet = book.active  # take active worksheet

    if 8 <= cur_hour and cur_hour < 20:  # find night shift names (II)
        for row in sheet.iter_rows(max_col=1):
            if row[0].value == "№ п/п":
                day_row_ind = row[0].row + 1
        for cell in sheet[day_row_ind]:
            if cell.value == cur_day:
                day_col_ind = cell.column
        for cell in sheet.iter_rows(min_col=day_col_ind - 1, max_col=day_col_ind - 1):
            if cell[0].value == "II":
                n_row = cell[0].row
                night_shift_names.append(sheet["C" + str(n_row - 1)].value)
        return (
            f"На смене в ночь были - {str(cur_day - 1) + '.'+ str(cur_month).zfill(2) + '.' + str(cur_year)}"
            + "\n"
            + "\n".join(night_shift_names)
        )
    else:  # find day shift names (I)
        for row in sheet.iter_rows(max_col=1):
            if row[0].value == "№ п/п":
                day_row_ind = row[0].row + 1
        if 20 <= cur_hour:  #  if current time before 00:00
            for cell in sheet[day_row_ind]:
                if cell.value == cur_day:
                    day_col_ind = cell.column
        else:  # if current time after 00:00
            for cell in sheet[day_row_ind]:
                if cell.value == cur_day - 1:
                    day_col_ind = cell.column
        for cell in sheet.iter_rows(min_col=day_col_ind, max_col=day_col_ind):
            if cell[0].value == "I":
                n_row = cell[0].row
                day_shift_names.append(sheet["C" + str(n_row - 1)].value)
        if cur_hour < 8:
            cur_day -= 1
        return (
            f"На смене в день были - {str(cur_day) +'.'+ str(cur_month).zfill(2) + '.' + str(cur_year)}"
            + "\n"
            + "\n".join(day_shift_names)
        )


def get_next_shift(schedule_file_name=F_NAME):
    """
    Parses work schedule and returns string of engeniers which will be on duty next shift .
    """

    cur_hour = int(datetime.datetime.now().strftime("%H"))
    cur_day = int(datetime.datetime.now().strftime("%d"))
    cur_month = int(datetime.datetime.now().strftime("%m"))
    cur_year = int(datetime.datetime.now().strftime("%Y"))

    day_shift_names = []
    night_shift_names = []

    book = openpyxl.load_workbook(schedule_file_name)  # open file
    sheet = book.active  # take active worksheet

    if 8 <= cur_hour and cur_hour < 20:  # find night shift names of current day (II)
        for row in sheet.iter_rows(max_col=1):
            if row[0].value == "№ п/п":
                day_row_ind = row[0].row + 1
        for cell in sheet[day_row_ind]:
            if cell.value == cur_day:
                day_col_ind = cell.column
        for cell in sheet.iter_rows(min_col=day_col_ind, max_col=day_col_ind):
            if cell[0].value == "II":
                n_row = cell[0].row
                night_shift_names.append(sheet["C" + str(n_row - 1)].value)
        return (
            f"На смене в ночь будут - {str(cur_day) +'.'+ str(cur_month).zfill(2) + '.' + str(cur_year)}"
            + "\n"
            + "\n".join(night_shift_names)
        )
    else:  # find  duty names of day shift next day (I)
        for row in sheet.iter_rows(max_col=1):
            if row[0].value == "№ п/п":
                day_row_ind = row[0].row + 1
        if 20 <= cur_hour:  #  if current time before 00:00
            for cell in sheet[day_row_ind]:
                if cell.value == cur_day + 1:
                    day_col_ind = cell.column
        else:  # if current time after 00:00
            for cell in sheet[day_row_ind]:
                if cell.value == cur_day:
                    day_col_ind = cell.column
        for cell in sheet.iter_rows(min_col=day_col_ind, max_col=day_col_ind):
            if cell[0].value == "I":
                n_row = cell[0].row
                day_shift_names.append(sheet["C" + str(n_row - 1)].value)
        if cur_hour >= 20:
            cur_day += 1
        return (
            f"На смене в день будут - {str(cur_day) +'.'+ str(cur_month).zfill(2) + '.' + str(cur_year)}"
            + "\n"
            + "\n".join(day_shift_names)
        )
